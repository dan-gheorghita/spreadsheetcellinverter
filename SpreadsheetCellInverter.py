#Spreadsheet Cell Inverter

import openpyxl

wb = openpyxl.load_workbook('Inverse.xlsx')

sheet = wb.active

wb_copy = openpyxl.Workbook()

sheet_copy = wb_copy.active

for col in range(sheet.max_column):
    for ro in range(sheet.max_row):
        sheet_copy.cell(row = col + 1, column = ro + 1).value = \
                       sheet.cell(row = ro + 1, column = col + 1).value

wb_copy.save('Inverse_copy.xlsx')
